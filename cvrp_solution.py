"""
IE 313 Assignment 3 - Capacitated Vehicle Routing Problem
Depot: Akhisar, Manisa
Trucks: Small (10t, 20lt/100km) and Large (15t, 40lt/100km)
Fuel: 65 TL/lt

NOTE ON VEHICLE SELECTION:
  Small truck: 10t capacity, 13 TL/km  (20 lt/100km × 65 TL/lt)
  Large truck: 15t capacity, 26 TL/km  (40 lt/100km × 65 TL/lt)

  Two small trucks vs one large truck:
    - 2 × Small: 20t capacity, 26 TL/km total  → more capacity, same cost
    - 1 × Large: 15t capacity, 26 TL/km         → less capacity, same cost

  Therefore, large trucks are STRICTLY DOMINATED by small trucks.
  The optimal solution will NEVER use a large truck.
  We formulate the model with only small trucks.
"""

import pyomo.environ as pyo
import openpyxl
import math

# =============================================================================
# 1. DATA LOADING
# =============================================================================

wb = openpyxl.load_workbook("ege-ilce.xlsx")

# Demand: use (city, district) tuples to handle duplicate district names
ws_demand = wb["Demand"]
demand_raw = {}
for row in ws_demand.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1]:
        demand_raw[(row[0], row[1])] = row[2] if row[2] else 0

DEPOT = ("MANİSA", "AKHİSAR")

# Customers: nodes with positive demand (excluding depot)
customers = sorted(
    [(city, dist) for (city, dist), dem in demand_raw.items() if dem > 0],
    key=lambda x: (x[0], x[1])
)

nodes = [DEPOT] + customers
n = len(nodes)
idx = {node: i for i, node in enumerate(nodes)}

total_demand = sum(demand_raw[c] for c in customers)
print(f"Depot: {DEPOT[1]} ({DEPOT[0]})")
print(f"Number of customers: {len(customers)}")
print(f"Total nodes (incl. depot): {n}")
print(f"Total demand: {total_demand:,} kg")

# Distance matrix: keyed by (city, district) tuples
ws_dist = wb["Distance"]
raw_dist = {}
for row in ws_dist.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1] and row[2] and row[3]:
        raw_dist[((row[0], row[1]), (row[2], row[3]))] = row[4]

def get_dist(ni, nj):
    if ni == nj:
        return 0
    d = raw_dist.get((ni, nj), raw_dist.get((nj, ni), None))
    if d is None:
        for (a, b), v in raw_dist.items():
            if a[1] == ni[1] and b[1] == nj[1]:
                return v
        print(f"WARNING: no distance found for {ni[1]} -> {nj[1]}, using 9999")
        return 9999
    return d

# Build full distance matrix
dist = {}
for i, ni in enumerate(nodes):
    for j, nj in enumerate(nodes):
        dist[i, j] = get_dist(ni, nj)

# Demands indexed by node index
dem = {0: 0}
for c in customers:
    dem[idx[c]] = demand_raw[c]

# =============================================================================
# PARAMETERS
# =============================================================================

CAP_SMALL = 10_000   # kg
CAP_LARGE = 15_000   # kg
FUEL_SMALL = 20      # lt/100km
FUEL_LARGE = 40      # lt/100km
FUEL_PRICE = 65      # TL/lt

COST_SMALL = FUEL_SMALL / 100 * FUEL_PRICE   # 13.0 TL/km
COST_LARGE = FUEL_LARGE / 100 * FUEL_PRICE   # 26.0 TL/km

print(f"\n--- Vehicle Cost Analysis ---")
print(f"Small truck: {CAP_SMALL/1000:.0f}t, {COST_SMALL:.1f} TL/km")
print(f"Large truck: {CAP_LARGE/1000:.0f}t, {COST_LARGE:.1f} TL/km")
print(f"2 × Small truck: {2*CAP_SMALL/1000:.0f}t, {2*COST_SMALL:.1f} TL/km")
print(f"=> Large truck is DOMINATED by 2 small trucks (same cost, more capacity)")
print(f"=> Using only small trucks in the model\n")

# Fleet: only small trucks (large trucks are dominated)
CAPACITY = CAP_SMALL
COST_PER_KM = COST_SMALL

# Use generous vehicle count — solver decides how many to use.
# Each unused vehicle stays at depot (cost = 0).
# OR-Tools solution showed 3 vehicles for unconstrained, 9 for time-constrained.
K = 6  # good balance: enough flexibility without exploding model size

vehicles = list(range(K))
N_range = range(n)
C_range = range(1, n)

print(f"Fleet: up to {K} small trucks (10t, {COST_PER_KM:.0f} TL/km)")
print(f"Total fleet capacity: {K * CAPACITY / 1000:.0f}t vs demand: {total_demand/1000:.1f}t")
print(f"Binary variables: ~{n * n * K:,}")


# =============================================================================
# 2. MODEL FORMULATION (CVRP - MTZ)
# =============================================================================

def build_model(time_constrained=False, n_vehicles=None):
    """
    Build Pyomo CVRP model with homogeneous fleet (small trucks only).

    Variables
    ---------
    x[i,j,k] in {0,1}  : vehicle k travels arc (i -> j)
    u[i,k]   >= 0       : MTZ cumulative load variable

    Objective:  min  COST_PER_KM * Σ dist[i,j] * x[i,j,k]

    Constraints:
    (1) Each customer visited exactly once (≤1 if time-constrained)
    (2) Flow conservation
    (3) Each vehicle leaves depot at most once
    (3b) Symmetry breaking
    (4) MTZ subtour elimination + capacity
    (5) Load bounds
    (6) Tour time ≤ 8h (if time-constrained)
    """
    veh = list(range(n_vehicles)) if n_vehicles else vehicles

    m = pyo.ConcreteModel()

    # --- Variables ---
    m.x = pyo.Var(N_range, N_range, veh, domain=pyo.Binary)
    m.u = pyo.Var(N_range, veh, domain=pyo.NonNegativeReals)

    # Fix self-loops and depot u
    for i in N_range:
        for k in veh:
            m.x[i, i, k].fix(0)
    for k in veh:
        m.u[0, k].fix(0)

    # --- Fix obviously infeasible arcs: depot->depot not needed, already 0 ---
    # Also fix arcs between very distant nodes to reduce search space
    # (a node pair that alone would exceed capacity can still be on same route,
    #  so we only prune self-loops which are already handled)

    # --- Objective ---
    m.obj = pyo.Objective(
        expr=sum(
            COST_PER_KM * dist[i, j] * m.x[i, j, k]
            for k in veh
            for i in N_range
            for j in N_range
            if i != j
        ),
        sense=pyo.minimize
    )

    # --- (1) Visit constraint ---
    def visit_rule(md, j):
        lhs = sum(md.x[i, j, k] for i in N_range if i != j for k in veh)
        if time_constrained:
            return lhs <= 1
        else:
            return lhs == 1
    m.visit = pyo.Constraint(C_range, rule=visit_rule)

    # --- (2) Flow conservation ---
    def flow_rule(md, i, k):
        out_ = sum(md.x[i, j, k] for j in N_range if j != i)
        in_  = sum(md.x[j, i, k] for j in N_range if j != i)
        return out_ == in_
    m.flow = pyo.Constraint(N_range, veh, rule=flow_rule)

    # --- (3) Each vehicle leaves depot at most once ---
    def depart_rule(md, k):
        return sum(md.x[0, j, k] for j in C_range) <= 1
    m.depart = pyo.Constraint(veh, rule=depart_rule)

    # --- (3b) Symmetry breaking ---
    def sym_break_rule(md, k):
        return (sum(md.x[0, j, k+1] for j in C_range)
                <= sum(md.x[0, j, k] for j in C_range))
    sym_pairs = [k for k in veh if k+1 in veh]
    m.sym_break = pyo.Constraint(sym_pairs, rule=sym_break_rule)

    # --- (4) MTZ subtour elimination ---
    def mtz_rule(md, i, j, k):
        if i == j or i == 0 or j == 0:
            return pyo.Constraint.Skip
        return md.u[i, k] + dem[j] - CAPACITY * (1 - md.x[i, j, k]) <= md.u[j, k]
    m.mtz = pyo.Constraint(N_range, N_range, veh, rule=mtz_rule)

    # --- (5) Load bounds ---
    def lb_rule(md, i, k):
        if i == 0:
            return pyo.Constraint.Skip
        return md.u[i, k] >= dem[i] * sum(md.x[j, i, k] for j in N_range if j != i)
    m.lb = pyo.Constraint(N_range, veh, rule=lb_rule)

    def ub_rule(md, i, k):
        return md.u[i, k] <= CAPACITY
    m.ub = pyo.Constraint(N_range, veh, rule=ub_rule)

    # --- (6) Time constraint (8 hours) ---
    if time_constrained:
        SPEED = 70          # km/h
        UNLOAD_RATE = 600   # kg/h
        TIME_LIMIT = 8.0    # hours

        def time_rule(md, k):
            drive_time  = sum(dist[i, j] / SPEED * md.x[i, j, k]
                              for i in N_range for j in N_range if i != j)
            unload_time = sum(dem[j] / UNLOAD_RATE * sum(md.x[i, j, k]
                              for i in N_range if i != j)
                              for j in C_range)
            return drive_time + unload_time <= TIME_LIMIT
        m.time_con = pyo.Constraint(veh, rule=time_rule)

    return m


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def extract_routes(model, veh_list=None):
    veh_list = veh_list or vehicles
    routes = {}
    for k in veh_list:
        if sum(pyo.value(model.x[0, j, k]) for j in C_range) < 0.5:
            continue
        route = [0]
        current = 0
        for _ in range(n):
            next_node = None
            for j in N_range:
                if j != current and pyo.value(model.x[current, j, k]) > 0.5:
                    next_node = j
                    break
            if next_node is None or next_node == 0:
                route.append(0)
                break
            route.append(next_node)
            current = next_node
        routes[k] = route
    return routes


def print_routes(routes, title="OPTIMAL ROUTES"):
    print("\n" + "="*70)
    print(title)
    print("="*70)
    total_cost = 0
    total_dist_sum = 0
    for k, route in routes.items():
        route_dist = sum(dist[route[i], route[i+1]] for i in range(len(route)-1))
        route_demand = sum(dem[j] for j in route if j != 0)
        route_cost = COST_PER_KM * route_dist

        node_names = [nodes[i][1] for i in route]
        print(f"\nVehicle {k+1} [Small (10t)]:")
        print(f"  Route: {' -> '.join(node_names)}")
        print(f"  Distance: {route_dist} km")
        print(f"  Load: {route_demand:,} kg / {CAPACITY:,} kg")
        print(f"  Fuel cost: {route_cost:,.1f} TL")
        total_cost += route_cost
        total_dist_sum += route_dist

    print(f"\n--- Summary ---")
    print(f"Vehicles used: {len(routes)}")
    print(f"Total distance: {total_dist_sum:,} km")
    print(f"Total fuel cost: {total_cost:,.1f} TL")
    return total_cost


def analyze_times(routes, title="TOUR TIME ANALYSIS"):
    SPEED = 70
    UNLOAD = 600
    print("\n" + "="*70)
    print(title)
    print("="*70)

    max_time = 0
    for k, route in routes.items():
        route_dist   = sum(dist[route[i], route[i+1]] for i in range(len(route)-1))
        route_demand = sum(dem[j] for j in route if j != 0)
        drive_h   = route_dist / SPEED
        unload_h  = route_demand / UNLOAD
        total_h   = drive_h + unload_h

        print(f"\nVehicle {k+1} [Small]:")
        print(f"  Driving time  : {drive_h:.2f} h  ({route_dist} km @ {SPEED} km/h)")
        print(f"  Unloading time: {unload_h:.2f} h  ({route_demand:,} kg @ 100 kg/10 min)")
        print(f"  Total tour time: {total_h:.2f} h  {'*** EXCEEDS 8h ***' if total_h > 8 else 'OK'}")
        max_time = max(max_time, total_h)

    print(f"\nLongest tour: {max_time:.2f} hours")
    return max_time


# =============================================================================
# 3-4. SOLVE CVRP (no time constraint)
# =============================================================================

print("\n" + "="*70)
print("SOLVING CVRP (no time constraint)...")
print("="*70)

model = build_model(time_constrained=False)

solver = pyo.SolverFactory("appsi_highs")
solver.options["mip_rel_gap"] = 0.0         # exact optimal, no gap
solver.options["time_limit"] = 25200.0      # 7 hour time limit
solver.options["threads"] = 0               # use all CPU cores

try:
    result = solver.solve(model, tee=True)
    print(f"\nSolver status: {result.solver.status}")
    print(f"Termination: {result.solver.termination_condition}")
    print(f"Objective value: {pyo.value(model.obj):,.1f} TL")
except RuntimeError as e:
    print(f"Solver error: {e}")
    print("No feasible solution found within time limit.")
    raise SystemExit(1)

routes = extract_routes(model)
total_cost = print_routes(routes, "PARTS 2-4: OPTIMAL ROUTES")


# =============================================================================
# 5. TOUR TIME CALCULATION
# =============================================================================

max_time = analyze_times(routes, "PART 5: TOUR TIME ANALYSIS")


# =============================================================================
# 6. IF ANY TOUR > 8h, ADD TIME CONSTRAINT AND RESOLVE
# =============================================================================

if max_time > 8.0:
    print("\n" + "="*70)
    print("PART 6: RESOLVING WITH 8-HOUR CONSTRAINT")
    print("=" * 70)
    print("Some tours exceed 8 hours. Adding time constraint...")
    print("Unsatisfied demand allowed if 8-hour shift is not sufficient.\n")

    # Need more vehicles when tours are shortened
    K2 = 10   # generous upper bound for time-constrained case
    print(f"Increasing fleet to max {K2} vehicles for time-constrained model")

    model2 = build_model(time_constrained=True, n_vehicles=K2)
    veh2 = list(range(K2))

    try:
        result2 = solver.solve(model2, tee=True)
        print(f"\nSolver status: {result2.solver.status}")
        print(f"Termination: {result2.solver.termination_condition}")
        print(f"Objective value: {pyo.value(model2.obj):,.1f} TL")
    except RuntimeError as e:
        print(f"Solver error: {e}")
        print("No feasible solution found within time limit.")
        raise SystemExit(1)

    routes2 = extract_routes(model2, veh2)
    total_cost2 = print_routes(routes2, "PART 6: TIME-CONSTRAINED ROUTES")

    # Verify tour times
    analyze_times(routes2, "PART 6: TOUR TIME VERIFICATION")

    # Unsatisfied demand
    served_nodes = set()
    for route in routes2.values():
        served_nodes.update(j for j in route if j != 0)

    unserved = [(nodes[j][1], nodes[j][0], dem[j])
                for j in C_range if j not in served_nodes]

    if unserved:
        print("\n--- Unsatisfied Demand ---")
        total_unserved = sum(d for _, _, d in unserved)
        for dist_name, city, d in sorted(unserved, key=lambda x: -x[2]):
            print(f"  {dist_name} ({city}): {d:,} kg")
        print(f"\nTotal unsatisfied demand: {total_unserved:,} kg "
              f"({100*total_unserved/total_demand:.1f}% of total)")
    else:
        print("\nAll demand satisfied within 8-hour constraint!")

    # Cost comparison
    print(f"\n--- Cost Comparison ---")
    print(f"Without time constraint: {total_cost:,.1f} TL ({len(routes)} vehicles)")
    print(f"With 8h constraint:      {total_cost2:,.1f} TL ({len(routes2)} vehicles)")
else:
    print("\nAll tours within 8 hours — no additional constraint needed.")
