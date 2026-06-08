"""
IE 313 Assignment 3 - Capacitated Vehicle Routing Problem (OR-Tools)
====================================================================
Depot: Akhisar, Manisa
Trucks: Small (10t, 20lt/100km) and Large (15t, 40lt/100km)
Fuel: 65 TL/lt

Vehicle Selection Analysis:
  Small truck: 10t capacity, cost = 20/100 × 65 = 13.0 TL/km
  Large truck: 15t capacity, cost = 40/100 × 65 = 26.0 TL/km

  Comparison: 2 small trucks vs 1 large truck
    2 × Small: 20t capacity, 2 × 13 = 26 TL/km  → MORE capacity, SAME cost
    1 × Large: 15t capacity, 26 TL/km             → LESS capacity, SAME cost

  => Large trucks are STRICTLY DOMINATED. Optimal solution uses only small trucks.

Solver: Google OR-Tools (Constraint Programming with metaheuristics)
"""

import openpyxl
import math
from ortools.constraint_solver import routing_enums_pb2, pywrapcp

# =============================================================================
# 1. DATA LOADING
# =============================================================================

print("=" * 70)
print("IE 313 ASSIGNMENT 3 - CVRP SOLUTION")
print("=" * 70)

wb = openpyxl.load_workbook("ege-ilce.xlsx")

# --- Demand ---
ws_demand = wb["Demand"]
demand_raw = {}
for row in ws_demand.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1]:
        demand_raw[(row[0], row[1])] = row[2] if row[2] else 0

DEPOT = ("MANİSA", "AKHİSAR")

# Customers: nodes with positive demand (excluding depot)
customers = sorted(
    [(city, dist) for (city, dist), dem in demand_raw.items() if dem > 0],
    key=lambda x: (x[0], x[1])
)

nodes = [DEPOT] + customers
n = len(nodes)
idx = {node: i for i, node in enumerate(nodes)}

total_demand = sum(demand_raw[c] for c in customers)
print(f"\nDepot: {DEPOT[1]} ({DEPOT[0]})")
print(f"Number of customers: {len(customers)}")
print(f"Total nodes (incl. depot): {n}")
print(f"Total demand: {total_demand:,} kg")

# --- Distance matrix ---
ws_dist = wb["Distance"]
raw_dist = {}
for row in ws_dist.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1] and row[2] and row[3]:
        raw_dist[((row[0], row[1]), (row[2], row[3]))] = row[4]


def get_dist(ni, nj):
    if ni == nj:
        return 0
    d = raw_dist.get((ni, nj), raw_dist.get((nj, ni), None))
    if d is None:
        for (a, b), v in raw_dist.items():
            if a[1] == ni[1] and b[1] == nj[1]:
                return v
        print(f"WARNING: no distance for {ni[1]} -> {nj[1]}, using 9999")
        return 9999
    return d


# Build integer distance matrix (OR-Tools requires integers)
dist_matrix = []
for i, ni in enumerate(nodes):
    row = []
    for j, nj in enumerate(nodes):
        row.append(int(get_dist(ni, nj)))
    dist_matrix.append(row)

# Demand list (indexed by node)
demands = [0] + [demand_raw[c] for c in customers]

# =============================================================================
# PARAMETERS
# =============================================================================

CAP_SMALL = 10_000   # kg
CAP_LARGE = 15_000   # kg
FUEL_SMALL = 20      # lt/100km
FUEL_LARGE = 40      # lt/100km
FUEL_PRICE = 65      # TL/lt

COST_SMALL = FUEL_SMALL / 100 * FUEL_PRICE   # 13.0 TL/km
COST_LARGE = FUEL_LARGE / 100 * FUEL_PRICE   # 26.0 TL/km

SPEED = 70       # km/h
UNLOAD = 600     # kg/h  (100 kg per 10 min = 600 kg/h)

print(f"\n--- Vehicle Cost Analysis ---")
print(f"Small truck: {CAP_SMALL/1000:.0f}t capacity, {COST_SMALL:.1f} TL/km")
print(f"Large truck: {CAP_LARGE/1000:.0f}t capacity, {COST_LARGE:.1f} TL/km")
print(f"2 × Small:   {2*CAP_SMALL/1000:.0f}t capacity, {2*COST_SMALL:.1f} TL/km")
print(f"=> Large truck is DOMINATED (same cost, less capacity than 2 small)")
print(f"=> Optimal fleet uses ONLY small trucks")

CAPACITY = CAP_SMALL
COST_PER_KM = COST_SMALL


# =============================================================================
# PART 1: MATHEMATICAL FORMULATION (printed for report)
# =============================================================================

print("\n" + "=" * 70)
print("PART 1: MATHEMATICAL FORMULATION")
print("=" * 70)
print("""
CVRP - Mixed Integer Programming Formulation
=============================================

Sets:
  N = {0, 1, ..., n-1}     set of all nodes (0 = depot)
  C = {1, ..., n-1}        set of customers
  K = {1, ..., K}          set of vehicles (all small trucks, 10t)

Parameters:
  d_ij  = distance from node i to node j (km)
  q_j   = demand at customer j (kg)
  Q     = vehicle capacity = 10,000 kg
  c     = cost per km = 13.0 TL/km  (20 lt/100km × 65 TL/lt)

Decision Variables:
  x_ijk ∈ {0,1}  = 1 if vehicle k travels from node i to node j
  u_ik  ≥ 0      = cumulative load of vehicle k after visiting node i (MTZ)

Objective:
  min  c × Σ_{k∈K} Σ_{i∈N} Σ_{j∈N, j≠i} d_ij × x_ijk

Subject to:
  (1) Σ_{i∈N, i≠j} Σ_{k∈K} x_ijk = 1          ∀ j ∈ C     (visit each customer exactly once)
  (2) Σ_{j∈N, j≠i} x_ijk = Σ_{j∈N, j≠i} x_jik  ∀ i∈N, k∈K  (flow conservation)
  (3) Σ_{j∈C} x_0jk ≤ 1                         ∀ k ∈ K     (each vehicle leaves depot at most once)
  (4) u_ik + q_j - Q(1 - x_ijk) ≤ u_jk          ∀ i,j∈C, k∈K (MTZ subtour elimination)
  (5) q_j ≤ u_jk ≤ Q                             ∀ j∈C, k∈K  (load bounds)
  (6) x_ijk ∈ {0,1}, u_ik ≥ 0
""")


# =============================================================================
# PARTS 2-4: SOLVE CVRP WITH OR-TOOLS
# =============================================================================

def solve_cvrp(num_vehicles, time_limit_per_route=None, search_time=30):
    """
    Solve CVRP using Google OR-Tools routing solver.

    Parameters
    ----------
    num_vehicles : int
        Number of available vehicles.
    time_limit_per_route : float or None
        If set, maximum tour time in hours per vehicle.
    search_time : int
        Solver time limit in seconds.

    Returns
    -------
    routes : list of lists
        Each route is a list of node indices [0, ..., 0].
    total_dist : int
        Total distance across all routes.
    """
    manager = pywrapcp.RoutingIndexManager(n, num_vehicles, 0)
    routing = pywrapcp.RoutingModel(manager)

    # --- Distance callback ---
    def distance_callback(from_index, to_index):
        from_node = manager.IndexToNode(from_index)
        to_node = manager.IndexToNode(to_index)
        return dist_matrix[from_node][to_node]

    transit_callback_index = routing.RegisterTransitCallback(distance_callback)
    routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)

    # --- Capacity constraint ---
    def demand_callback(from_index):
        from_node = manager.IndexToNode(from_index)
        return demands[from_node]

    demand_callback_index = routing.RegisterUnaryTransitCallback(demand_callback)
    routing.AddDimensionWithVehicleCapacity(
        demand_callback_index,
        0,                                      # no slack
        [CAPACITY] * num_vehicles,              # vehicle capacities
        True,                                   # start cumul to zero
        "Capacity"
    )

    # --- Time constraint (if applicable) ---
    if time_limit_per_route is not None:
        # Time = driving time + unloading time
        # We work in minutes (integer) for OR-Tools
        def time_callback(from_index, to_index):
            from_node = manager.IndexToNode(from_index)
            to_node = manager.IndexToNode(to_index)
            # Driving time (minutes)
            drive_min = dist_matrix[from_node][to_node] / SPEED * 60
            # Unloading time at destination (minutes)
            unload_min = demands[to_node] / UNLOAD * 60
            return int(round(drive_min + unload_min))

        time_callback_index = routing.RegisterTransitCallback(time_callback)
        time_limit_min = int(time_limit_per_route * 60)
        routing.AddDimension(
            time_callback_index,
            0,                      # no slack
            time_limit_min,         # max time per vehicle
            True,                   # start cumul to zero
            "Time"
        )

    # --- Allow dropping nodes (for time-constrained version) ---
    if time_limit_per_route is not None:
        # Large penalty for dropping nodes (but allows it if infeasible)
        penalty = 100_000
        for node_idx in range(1, n):
            routing.AddDisjunction([manager.NodeToIndex(node_idx)], penalty)

    # --- Search parameters ---
    search_params = pywrapcp.DefaultRoutingSearchParameters()
    search_params.first_solution_strategy = (
        routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
    )
    search_params.local_search_metaheuristic = (
        routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    )
    search_params.time_limit.seconds = search_time

    # --- Solve ---
    solution = routing.SolveWithParameters(search_params)

    if solution is None:
        print("ERROR: No solution found!")
        return None, None

    # --- Extract routes ---
    routes = []
    total_dist = 0
    for vehicle_id in range(num_vehicles):
        route = []
        index = routing.Start(vehicle_id)
        while not routing.IsEnd(index):
            node = manager.IndexToNode(index)
            route.append(node)
            index = solution.Value(routing.NextVar(index))
        route.append(0)  # return to depot

        if len(route) > 2:  # skip empty routes (depot -> depot)
            routes.append(route)
            route_dist = sum(dist_matrix[route[i]][route[i+1]]
                             for i in range(len(route)-1))
            total_dist += route_dist

    return routes, total_dist


def print_routes(routes, title="OPTIMAL ROUTES"):
    """Print detailed route information."""
    print("\n" + "=" * 70)
    print(title)
    print("=" * 70)

    total_cost = 0
    total_distance = 0

    for v, route in enumerate(routes):
        route_dist = sum(dist_matrix[route[i]][route[i+1]]
                         for i in range(len(route)-1))
        route_demand = sum(demands[j] for j in route if j != 0)
        route_cost = COST_PER_KM * route_dist

        node_names = [nodes[i][1] for i in route]
        print(f"\nVehicle {v+1} [Small (10t)]:")
        print(f"  Route: {' -> '.join(node_names)}")
        print(f"  Distance: {route_dist} km")
        print(f"  Load: {route_demand:,} kg / {CAPACITY:,} kg")
        print(f"  Fuel cost: {route_cost:,.1f} TL")

        total_cost += route_cost
        total_distance += route_dist

    print(f"\n--- Summary ---")
    print(f"Vehicles used: {len(routes)}")
    print(f"Total distance: {total_distance:,} km")
    print(f"Total fuel cost: {total_cost:,.1f} TL")
    return total_cost


def analyze_tour_times(routes, title="TOUR TIME ANALYSIS"):
    """Part 5: Calculate and display tour times."""
    print("\n" + "=" * 70)
    print(title)
    print("=" * 70)

    max_time = 0
    any_exceeds = False

    for v, route in enumerate(routes):
        route_dist = sum(dist_matrix[route[i]][route[i+1]]
                         for i in range(len(route)-1))
        route_demand = sum(demands[j] for j in route if j != 0)

        drive_h = route_dist / SPEED
        unload_h = route_demand / UNLOAD
        total_h = drive_h + unload_h
        exceeds = total_h > 8.0

        print(f"\nVehicle {v+1} [Small]:")
        print(f"  Driving time  : {drive_h:.2f} h  ({route_dist} km @ {SPEED} km/h)")
        print(f"  Unloading time: {unload_h:.2f} h  ({route_demand:,} kg @ 100 kg/10 min)")
        print(f"  Total tour time: {total_h:.2f} h  {'*** EXCEEDS 8h ***' if exceeds else 'OK'}")

        max_time = max(max_time, total_h)
        if exceeds:
            any_exceeds = True

    print(f"\nLongest tour: {max_time:.2f} hours")
    return max_time, any_exceeds


# =============================================================================
# SOLVE PART 2-4: CVRP without time constraint
# =============================================================================

print("\n" + "=" * 70)
print("PARTS 2-4: SOLVING CVRP (no time constraint)")
print("=" * 70)

# Use enough vehicles — let the solver decide how many to actually use
num_vehicles = max(3, math.ceil(total_demand / CAPACITY)) + 3
print(f"Max vehicles available: {num_vehicles}")
print(f"Solving with OR-Tools (30 second search)...")

routes, total_dist = solve_cvrp(num_vehicles=num_vehicles, search_time=30)

if routes:
    total_cost = print_routes(routes, "PARTS 2-4: OPTIMAL ROUTES")

    # =================================================================
    # PART 5: Tour time analysis
    # =================================================================
    max_time, any_exceeds = analyze_tour_times(routes, "PART 5: TOUR TIME ANALYSIS")

    # =================================================================
    # PART 6: 8-hour constraint
    # =================================================================
    if any_exceeds:
        print("\n" + "=" * 70)
        print("PART 6: RESOLVING WITH 8-HOUR CONSTRAINT")
        print("=" * 70)
        print("Some tours exceed 8 hours. Adding time constraint...")
        print("Unsatisfied demand allowed if 8-hour shift is not sufficient.\n")

        # More vehicles may be needed when tours are shortened
        num_vehicles_tc = num_vehicles + 3
        print(f"Max vehicles available: {num_vehicles_tc}")
        print(f"Solving with 8-hour constraint...")

        routes2, total_dist2 = solve_cvrp(
            num_vehicles=num_vehicles_tc,
            time_limit_per_route=8.0,
            search_time=30
        )

        if routes2:
            total_cost2 = print_routes(routes2, "PART 6: TIME-CONSTRAINED ROUTES")

            # Tour time verification
            analyze_tour_times(routes2, "PART 6: TOUR TIME VERIFICATION")

            # Unsatisfied demand analysis
            served_nodes = set()
            for route in routes2:
                served_nodes.update(j for j in route if j != 0)

            unserved = [(nodes[j][1], nodes[j][0], demands[j])
                        for j in range(1, n) if j not in served_nodes]

            if unserved:
                print("\n--- Unsatisfied Demand ---")
                total_unserved = sum(d for _, _, d in unserved)
                for dist_name, city, d in sorted(unserved, key=lambda x: -x[2]):
                    print(f"  {dist_name} ({city}): {d:,} kg")
                print(f"\nTotal unsatisfied demand: {total_unserved:,} kg "
                      f"({100*total_unserved/total_demand:.1f}% of total)")
            else:
                print("\nAll demand satisfied within 8-hour constraint!")

            # Cost comparison
            print(f"\n--- Cost Comparison ---")
            print(f"Without time constraint: {total_cost:,.1f} TL ({len(routes)} vehicles)")
            print(f"With 8h constraint:      {total_cost2:,.1f} TL ({len(routes2)} vehicles)")
            if total_cost2 > total_cost:
                print(f"Cost increase:           {total_cost2 - total_cost:,.1f} TL "
                      f"({(total_cost2 - total_cost)/total_cost*100:.1f}%)")
    else:
        print("\n" + "=" * 70)
        print("PART 6: 8-HOUR CONSTRAINT")
        print("=" * 70)
        print("All tours are within 8 hours — no additional constraint needed.")
        print("The original solution already satisfies the 8-hour shift requirement.")
